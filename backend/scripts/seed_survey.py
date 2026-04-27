"""
Seed script: imports TFA Training Landscape Study survey from Excel workbook.

Usage (from backend/ directory):
    python scripts/seed_survey.py --xlsx ../04262027_TFA\ Strategy\ Refresh_*.xlsx
"""
import argparse
import logging
import os
import sys
from dataclasses import dataclass, field

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.core.database import SessionLocal
from app.models.survey import QuestionType, RespondentRole, Survey
from app.repositories.survey_repository import SurveyRepository

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

SURVEY_SLUG = "tfa-training-landscape-2025"

# Map from Excel format strings to QuestionType enum values
FORMAT_MAP = {
    "open text": QuestionType.OPEN_TEXT,
    "open text + interviewer code": QuestionType.OPEN_TEXT,
    "pre-filled answers (optional: open text)": QuestionType.SINGLE_CHOICE,
}


@dataclass
class QuestionData:
    number: str
    module: str
    objective: str
    text_en: str
    question_type: QuestionType
    options_en: list[str]
    has_open_text: bool
    open_text_label_en: str | None
    visible_to_ceo: bool
    visible_to_chro: bool
    visible_to_ld: bool
    section_key: str = ""


@dataclass
class SectionData:
    key: str
    title_en: str
    order: int
    questions: list[QuestionData] = field(default_factory=list)


def parse_bool(val: object) -> bool:
    if val is None:
        return False
    return str(val).strip().lower() in ("yes", "y", "true", "1")


def clean(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip()


def detect_format(fmt_str: str) -> tuple[QuestionType, bool, str | None]:
    """Returns (question_type, has_open_text_option, open_text_label)."""
    normalized = fmt_str.lower().strip()
    qt = FORMAT_MAP.get(normalized, QuestionType.OPEN_TEXT)
    has_open = qt == QuestionType.SINGLE_CHOICE
    open_label = "Other (please specify)" if has_open else None
    return qt, has_open, open_label


def parse_overview_sheet(ws) -> list[QuestionData]:
    questions: list[QuestionData] = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue

        cells = list(row)
        # Header detection: row starts with None and '#' in column 1
        if not header_found:
            if len(cells) > 1 and cells[1] == "#":
                header_found = True
            continue

        num_raw = cells[1]
        if num_raw is None:
            continue
        if isinstance(num_raw, str) and num_raw.upper().startswith("SECTION"):
            continue

        num = clean(num_raw)
        module = clean(cells[2])
        objective = clean(cells[3])
        text_en = clean(cells[4])
        fmt = clean(cells[5])

        if not text_en:
            continue

        options_raw = [clean(cells[i]) for i in range(6, 11) if i < len(cells)]
        options_en = [o for o in options_raw if o and o != "-"]

        open_text_col = clean(cells[11]) if len(cells) > 11 else ""
        has_open_from_col = bool(open_text_col)
        open_label = open_text_col if open_text_col else None

        qt, has_open, _default_label = detect_format(fmt)
        if has_open_from_col:
            has_open = True

        ceo = parse_bool(cells[13]) if len(cells) > 13 else False
        chro = parse_bool(cells[14]) if len(cells) > 14 else False
        ld_raw = cells[15] if len(cells) > 15 else None
        ld = parse_bool(ld_raw)

        questions.append(QuestionData(
            number=num,
            module=module,
            objective=objective,
            text_en=text_en,
            question_type=qt,
            options_en=options_en,
            has_open_text=has_open,
            open_text_label_en=open_label,
            visible_to_ceo=ceo,
            visible_to_chro=chro,
            visible_to_ld=ld,
        ))

    return questions


def parse_role_sheet(ws) -> dict[str, list[dict]]:
    """Returns ordered list of {section_key, section_title, question_number} per sheet."""
    result: dict[str, list] = {}
    current_section_key = ""
    current_section_title = ""
    order = 0

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not any(c is not None for c in cells):
            continue

        # Check if this is a section header row
        first_cell = clean(cells[0]) if cells[0] is not None else ""
        if first_cell.upper().startswith("SECTION"):
            current_section_key = first_cell.upper().replace(" ", "_").replace(":", "").replace("/", "_")
            current_section_title = first_cell
            order += 1
            result[current_section_key] = []
            continue

        # Skip column header row
        if first_cell == "#":
            continue

        num_raw = cells[0]
        if num_raw is None:
            continue
        q_num = clean(num_raw)

        # Append question number to current section
        if current_section_key and q_num:
            result[current_section_key].append({
                "section_key": current_section_key,
                "section_title": current_section_title,
                "question_number": q_num,
                "order": order,
            })

    return result


def group_into_sections(role_sheet_data: dict[str, list]) -> list[SectionData]:
    """Build SectionData list from role sheet parse result."""
    sections: list[SectionData] = []
    seen = set()
    order = 0
    for section_key, q_entries in role_sheet_data.items():
        if not q_entries:
            continue
        title = q_entries[0]["section_title"] if q_entries else section_key
        # Convert SECTION_0:_INTRODUCTION -> section_0_introduction
        key = section_key.lower().replace(" ", "_")
        sections.append(SectionData(key=key, title_en=title, order=order))
        order += 1
    return sections


def build_survey_structure(overview_questions: list[QuestionData], role_sheets: dict) -> dict:
    """Combine overview questions with per-role section info into a complete structure."""

    # Build lookup: question_number -> QuestionData
    q_lookup: dict[str, QuestionData] = {}
    for q in overview_questions:
        q_lookup[str(q.number)] = q

    # Collect all unique sections across all role sheets
    all_sections: dict[str, SectionData] = {}
    section_order = 0

    for role_name, sheet_data in role_sheets.items():
        for section_key, q_entries in sheet_data.items():
            if section_key not in all_sections:
                title = q_entries[0]["section_title"] if q_entries else section_key
                sk = section_key.lower().replace(" ", "_").replace(":", "").replace("/", "_").replace("__", "_")
                all_sections[section_key] = SectionData(key=sk, title_en=title, order=section_order)
                section_order += 1

            # Assign section_key to questions
            for entry in q_entries:
                q_num = entry["question_number"]
                if q_num in q_lookup:
                    q = q_lookup[q_num]
                    if not q.section_key:
                        q.section_key = all_sections[section_key].key

    return {
        "sections": list(all_sections.values()),
        "questions": list(q_lookup.values()),
    }


def seed(xlsx_path: str) -> None:
    logger.info("[Seed] Loading workbook: %s", xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)

    overview_ws = wb["0. Overview Questions (all)"]
    ceo_ws = wb["1. CEO Survey"]
    chro_ws = wb["2. CHRO Survey"]
    ld_ws = wb["3. L&D Survey"]

    overview_questions = parse_overview_sheet(overview_ws)
    logger.info("[Seed] Parsed %d questions from overview", len(overview_questions))

    ceo_data = parse_role_sheet(ceo_ws)
    chro_data = parse_role_sheet(chro_ws)
    ld_data = parse_role_sheet(ld_ws)

    structure = build_survey_structure(
        overview_questions,
        {"ceo": ceo_data, "chro": chro_data, "ld": ld_data},
    )

    db = SessionLocal()
    try:
        repo = SurveyRepository(db)

        # Check if survey already exists
        existing = repo.get_by_slug(SURVEY_SLUG)
        if existing:
            logger.info("[Seed] Survey already exists (slug=%s), skipping", SURVEY_SLUG)
            return

        logger.info("[Seed] Creating survey: %s", SURVEY_SLUG)
        survey = repo.create(slug=SURVEY_SLUG, is_active=True, is_fs_only=False)

        repo.upsert_translation(
            survey.id, "en",
            title="The Financial Academy Training Landscape Study",
            description=(
                "This survey is part of the Financial Academy's Strategy Refresh initiative. "
                "Your responses will help shape the future of financial services training in Saudi Arabia."
            ),
            instructions=(
                "Please answer all questions honestly. Your responses are confidential and will only be "
                "used in aggregate to inform The Financial Academy's strategy."
            ),
            thank_you_message=(
                "Thank you for completing this survey. Your input is invaluable to the Financial Academy."
            ),
        )

        repo.upsert_translation(
            survey.id, "ar",
            title="دراسة مشهد التدريب في الأكاديمية المالية",
            description=(
                "يُعدّ هذا الاستطلاع جزءًا من مبادرة تحديث استراتيجية الأكاديمية المالية. "
                "ستُساهم إجاباتك في رسم مستقبل تدريب الخدمات المالية في المملكة العربية السعودية."
            ),
            instructions=(
                "يُرجى الإجابة على جميع الأسئلة بصدق. إجاباتك سرية ولن تُستخدم إلا بشكل إجمالي "
                "لدعم استراتيجية الأكاديمية المالية."
            ),
            thank_you_message=(
                "شكرًا لإتمامك هذا الاستطلاع. مشاركتك ذات قيمة لا تُقدّر للأكاديمية المالية."
            ),
        )

        # Build sections by grouping questions
        sections_map = _derive_sections_from_sheets(ceo_ws, chro_ws, ld_ws, overview_questions)

        section_id_map: dict[str, int] = {}
        for sec_order, (sec_key, sec_info) in enumerate(sections_map.items()):
            section = repo.create_section(
                survey_id=survey.id,
                section_key=sec_key,
                display_order=sec_order,
            )
            repo.upsert_section_translation(section.id, "en", title=sec_info["title_en"])
            repo.upsert_section_translation(section.id, "ar", title=sec_info.get("title_ar", sec_info["title_en"]))
            section_id_map[sec_key] = section.id

        # Create questions
        for q_idx, q in enumerate(overview_questions):
            sec_key = q.section_key or "section_unknown"
            if sec_key not in section_id_map:
                # Create a fallback section
                sec = repo.create_section(
                    survey_id=survey.id,
                    section_key=sec_key,
                    display_order=len(section_id_map),
                )
                repo.upsert_section_translation(sec.id, "en", title=sec_key.replace("_", " ").title())
                repo.upsert_section_translation(sec.id, "ar", title=sec_key.replace("_", " ").title())
                section_id_map[sec_key] = sec.id

            question = repo.create_question(
                section_id=section_id_map[sec_key],
                question_key=f"q_{q.number}",
                question_type=q.question_type,
                display_order=q_idx,
                is_required=True,
                has_open_text_option=q.has_open_text,
                open_text_label_en=q.open_text_label_en or "Other (please specify)",
                open_text_label_ar="أخرى (يرجى التحديد)",
                module=q.module,
            )

            repo.upsert_question_translation(question.id, "en", text=q.text_en)
            # Arabic placeholder — translators will fill these in
            repo.upsert_question_translation(question.id, "ar", text=q.text_en)

            # Create options
            for opt_idx, opt_text in enumerate(q.options_en):
                opt_key = f"opt_{opt_idx + 1}"
                option = repo.create_option(
                    question_id=question.id,
                    option_key=opt_key,
                    display_order=opt_idx,
                )
                repo.upsert_option_translation(option.id, "en", opt_text)
                repo.upsert_option_translation(option.id, "ar", opt_text)

            # Set visibility rules
            roles = []
            if q.visible_to_ceo:
                roles.append(RespondentRole.CEO)
            if q.visible_to_chro:
                roles.append(RespondentRole.CHRO)
            if q.visible_to_ld:
                roles.append(RespondentRole.LD)
            repo.set_visibility_rules(question.id, roles)

        db.commit()
        logger.info("[Seed] Survey seeded successfully with %d questions", len(overview_questions))

    except Exception as exc:
        db.rollback()
        logger.error("[Seed] Failed: %s", str(exc))
        raise
    finally:
        db.close()


def _derive_sections_from_sheets(ceo_ws, chro_ws, ld_ws, questions: list[QuestionData]) -> dict:
    """Extract unique sections by reading section header rows from all three sheets."""

    SECTION_TRANSLATIONS_AR = {
        "INTRODUCTION": "المقدمة",
        "ORGANIZATIONAL CONTEXT": "السياق التنظيمي",
        "INSTITUTIONAL CONTEXT": "السياق المؤسسي",
        "L&D STRATEGY AND BUDGET": "استراتيجية التعلم والتطوير والميزانية",
        "L&D FUNCTION CONTEXT": "سياق وظيفة التعلم والتطوير",
        "STRATEGIC L&D POSITIONING": "التموضع الاستراتيجي للتعلم والتطوير",
        "ECONOMIC CONTEXT": "السياق الاقتصادي",
        "CURRENT PROVIDER LANDSCAPE (UNAIDED)": "مشهد مزودي التدريب الحالي",
        "PROVIDER SELECTION AND DELIVERY": "اختيار المزود وأسلوب التقديم",
        "THE FINANCIAL ACADEMY PERCEPTION": "تصورات الأكاديمية المالية",
        "DELIVERY AND FORMAT": "أسلوب التقديم والتنسيق",
        "DEMAND, OUTCOMES": "الطلب والنتائج",
        "DEMAND THEMES": "محاور الطلب",
        "CERTIFICATIONS": "الشهادات المهنية",
        "DIGITAL AND OUTCOMES": "الرقمنة والنتائج",
        "PROCUREMENT": "المشتريات",
        "MARKET VIEW": "نظرة السوق",
        "FORWARD-LOOKING": "التوقعات المستقبلية",
    }

    sections_ordered: dict[str, dict] = {}

    for ws in [ceo_ws, chro_ws, ld_ws]:
        order = 0
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            if not any(c is not None for c in cells):
                continue
            first = clean(cells[0])
            if first.upper().startswith("SECTION"):
                # Extract title part after "SECTION N:"
                parts = first.split(":", 1)
                title_en = parts[1].strip() if len(parts) > 1 else first.strip()
                # Normalize key
                sec_key = f"section_{order}"
                if sec_key not in sections_ordered:
                    title_upper = title_en.upper()
                    # Find best Arabic translation
                    title_ar = None
                    for key_fragment, ar_title in SECTION_TRANSLATIONS_AR.items():
                        if key_fragment in title_upper:
                            title_ar = ar_title
                            break

                    sections_ordered[sec_key] = {
                        "title_en": title_en,
                        "title_ar": title_ar or title_en,
                        "order": len(sections_ordered),
                    }
                order += 1

    # Assign section keys to questions based on order in overview
    current_section_keys = list(sections_ordered.keys())
    # We assign based on the order questions appear
    if questions:
        # Distribute questions across sections by re-reading the overview structure
        # Use CEO sheet as reference for section assignment
        _assign_sections_from_sheet(ceo_ws, questions, current_section_keys)

    return sections_ordered


def _assign_sections_from_sheet(ws, questions: list[QuestionData], section_keys: list[str]) -> None:
    """Walk through a sheet and assign section_key to questions by matching question numbers."""
    q_lookup = {clean(q.number): q for q in questions}

    current_section_idx = -1
    section_key = ""

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not any(c is not None for c in cells):
            continue
        first = clean(cells[0])
        if first.upper().startswith("SECTION"):
            current_section_idx += 1
            if current_section_idx < len(section_keys):
                section_key = section_keys[current_section_idx]
            continue

        if first in ("#", ""):
            continue

        num = clean(cells[0])
        if num in q_lookup and not q_lookup[num].section_key:
            q_lookup[num].section_key = section_key

    # For any still unassigned, use the last valid section
    last_key = section_keys[-1] if section_keys else "section_0"
    for q in questions:
        if not q.section_key:
            q.section_key = last_key


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Seed TFA survey from Excel workbook")
    parser.add_argument(
        "--xlsx",
        default="../04262027_TFA Strategy Refresh_Questionnaire Training Landscape Study_v3.xlsx",
        help="Path to the Excel workbook",
    )
    args = parser.parse_args()
    seed(args.xlsx)
