from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy.orm import Session

from app.api.deps import get_current_admin, require_permission
from app.core.database import get_db
from app.models.admin import AdminUser
from app.repositories.admin_repository import AdminRepository
from app.repositories.group_registration_repository import GroupRegistrationRepository
from app.schemas.group_registration import (
    GroupRegistrationConfigCreate,
    GroupRegistrationConfigOut,
    GroupRegistrationConfigUpdate,
    GroupRegistrationOut,
    TrainingCourseCreate,
    TrainingCourseOut,
    TrainingCourseUpdate,
)

router = APIRouter(prefix="/group-registration", tags=["admin-group-registration"])

PERM = "surveys.group_registration.manage"


# ── Submissions ───────────────────────────────────────────────────────────────

@router.get("", response_model=dict)
def list_registrations(
    skip:              int = Query(0, ge=0),
    limit:             int = Query(50, ge=1, le=200),
    status_filter:     str | None = Query(None, alias="status"),
    organization_name: str | None = Query(None),
    db:    Session  = Depends(get_db),
    admin: AdminUser = Depends(require_permission(PERM)),
):
    repo = GroupRegistrationRepository(db)
    items, total = repo.list_all(
        skip=skip, limit=limit,
        status=status_filter,
        organization_name=organization_name,
    )
    return {
        "total": total,
        "items": [
            {
                "id":               r.id,
                "reference_number": r.reference_number,
                "organization_name":r.organization_name,
                "focal_point_name": r.focal_point_name,
                "email":            r.email,
                "nomination_count": len(r.nominations or []),
                "status":           r.status,
                "submitted_at":     r.submitted_at.isoformat() if r.submitted_at else None,
                "created_at":       r.created_at.isoformat(),
            }
            for r in items
        ],
    }


@router.get("/{reg_id}", response_model=GroupRegistrationOut)
def get_registration(
    reg_id: int,
    db:    Session  = Depends(get_db),
    _:     AdminUser = Depends(require_permission(PERM)),
):
    reg = GroupRegistrationRepository(db).get_by_id(reg_id)
    if not reg:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registration not found")
    return reg


@router.get("/export/xlsx")
def export_xlsx(
    db:    Session  = Depends(get_db),
    admin: AdminUser = Depends(require_permission(PERM)),
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    repo  = GroupRegistrationRepository(db)
    items, _ = repo.list_all(skip=0, limit=10_000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Group Registrations"

    header_fill = PatternFill("solid", fgColor="1B3A5C")
    header_font = Font(color="FFFFFF", bold=True)

    headers = [
        "Reference", "Organization", "Department", "Focal Point", "Position",
        "Email", "Mobile", "Sectors", "Functional Areas",
        "# Nominations", "Special Requests", "Submitted By", "PDPL", "Status", "Submitted At",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in items:
        ws.append([
            r.reference_number,
            r.organization_name,
            r.department or "",
            r.focal_point_name,
            r.focal_point_position or "",
            r.email,
            r.mobile or "",
            ", ".join(r.selected_sectors or []),
            ", ".join(r.selected_functional_areas or []),
            len(r.nominations or []),
            r.special_requests or "",
            r.submitted_by or "",
            "Yes" if r.pdpl_authorized else "No",
            r.status,
            r.submitted_at.isoformat() if r.submitted_at else "",
        ])

    # Nominations detail sheet
    ws2 = wb.create_sheet("Nominations Detail")
    ws2.append(["Reference", "Org", "Row #", "Sector", "Functional Area",
                "Course Code", "Course Title", "Delivery Mode", "Quarters", "Nominations"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    for r in items:
        for i, nom in enumerate(r.nominations or [], start=1):
            ws2.append([
                r.reference_number,
                r.organization_name,
                i,
                nom.get("sector", ""),
                nom.get("functional_area", ""),
                nom.get("course_code", ""),
                nom.get("course_title", ""),
                nom.get("delivery_mode", ""),
                ", ".join(f"Q{q}" for q in nom.get("preferred_quarters", [])),
                nom.get("num_nominations", ""),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=group_registrations.xlsx"},
    )


# ── Form configs ─────────────────────────────────────────────────────────────

@router.get("/configs", response_model=list[GroupRegistrationConfigOut])
def list_configs(
    db: Session  = Depends(get_db),
    _:  AdminUser = Depends(require_permission(PERM)),
):
    return GroupRegistrationRepository(db).list_configs()


@router.post("/configs", response_model=GroupRegistrationConfigOut, status_code=status.HTTP_201_CREATED)
def create_config(
    body:  GroupRegistrationConfigCreate,
    db:    Session  = Depends(get_db),
    admin: AdminUser = Depends(require_permission(PERM)),
):
    repo = GroupRegistrationRepository(db)
    if repo.get_config_by_slug(body.slug):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Slug already exists")
    cfg = repo.create_config(body.model_dump())
    AdminRepository(db).log_action(admin.id, "config_created", "group_reg_config", body.slug)
    db.commit()
    db.refresh(cfg)
    return cfg


@router.put("/configs/{config_id}", response_model=GroupRegistrationConfigOut)
def update_config(
    config_id: int,
    body:      GroupRegistrationConfigUpdate,
    db:        Session  = Depends(get_db),
    admin:     AdminUser = Depends(require_permission(PERM)),
):
    repo = GroupRegistrationRepository(db)
    cfg  = repo.get_config(config_id)
    if not cfg:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Config not found")
    repo.update_config(cfg, body.model_dump(exclude_none=True))
    AdminRepository(db).log_action(admin.id, "config_updated", "group_reg_config", str(config_id))
    db.commit()
    db.refresh(cfg)
    return cfg


@router.delete("/configs/{config_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_config(
    config_id: int,
    db:        Session  = Depends(get_db),
    admin:     AdminUser = Depends(require_permission(PERM)),
):
    repo = GroupRegistrationRepository(db)
    cfg  = repo.get_config(config_id)
    if not cfg:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Config not found")
    db.delete(cfg)
    AdminRepository(db).log_action(admin.id, "config_deleted", "group_reg_config", str(config_id))
    db.commit()


# ── Course catalog management ─────────────────────────────────────────────────

@router.get("/courses", response_model=list[TrainingCourseOut])
def list_courses(
    db: Session  = Depends(get_db),
    _:  AdminUser = Depends(require_permission(PERM)),
):
    return GroupRegistrationRepository(db).list_courses(active_only=False)


@router.post("/courses", response_model=TrainingCourseOut, status_code=status.HTTP_201_CREATED)
def create_course(
    body:  TrainingCourseCreate,
    db:    Session  = Depends(get_db),
    admin: AdminUser = Depends(require_permission(PERM)),
):
    repo = GroupRegistrationRepository(db)
    if repo.get_course_by_code(body.course_code):
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Course code already exists")
    course = repo.create_course(body.model_dump())
    AdminRepository(db).log_action(admin.id, "course_created", "training_course", course.course_code)
    db.commit()
    db.refresh(course)
    return course


@router.put("/courses/{course_id}", response_model=TrainingCourseOut)
def update_course(
    course_id: int,
    body:      TrainingCourseUpdate,
    db:        Session  = Depends(get_db),
    admin:     AdminUser = Depends(require_permission(PERM)),
):
    repo   = GroupRegistrationRepository(db)
    course = repo.get_course(course_id)
    if not course:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found")
    repo.update_course(course, body.model_dump(exclude_none=True))
    AdminRepository(db).log_action(admin.id, "course_updated", "training_course", str(course_id))
    db.commit()
    db.refresh(course)
    return course


@router.delete("/courses/{course_id}", status_code=status.HTTP_204_NO_CONTENT)
def deactivate_course(
    course_id: int,
    db:        Session  = Depends(get_db),
    admin:     AdminUser = Depends(require_permission(PERM)),
):
    repo   = GroupRegistrationRepository(db)
    course = repo.get_course(course_id)
    if not course:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found")
    course.is_active = False
    AdminRepository(db).log_action(admin.id, "course_deactivated", "training_course", str(course_id))
    db.commit()
