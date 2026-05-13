import io
import csv
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import select

from app.models.submission import Submission, Answer
from app.models.survey import Question, QuestionOption


# ── Translation / answer helpers ─────────────────────────────────────────────

def _t(translations, lang: str) -> str:
    for t in translations:
        if t.language_code == lang:
            return t.text or ""
    for t in translations:
        if t.language_code == "en":
            return t.text or ""
    return translations[0].text if translations else ""


def _answer_text(answer: Answer, lang: str) -> str:
    q = answer.question
    if not q:
        return ""
    parts: list[str] = []
    if answer.selected_option_keys:
        opt_map = {o.option_key: o for o in q.options}
        for key in answer.selected_option_keys:
            opt = opt_map.get(key)
            parts.append(_t(opt.translations, lang) if opt else key)
    if answer.open_text_value:
        parts.append(answer.open_text_value)
    if answer.numeric_value is not None and not parts:
        parts.append(str(answer.numeric_value))
    return " | ".join(parts)


# ── Data loading ──────────────────────────────────────────────────────────────

def _load_submissions(
    db:        Session,
    survey_id: int | None,
    org_id:    int | None,
    role:      str | None,
    status:    str | None,
) -> list[Submission]:
    q = (
        select(Submission)
        .options(
            joinedload(Submission.organization),
            selectinload(Submission.answers)
            .selectinload(Answer.question)
            .selectinload(Question.translations),
            selectinload(Submission.answers)
            .selectinload(Answer.question)
            .selectinload(Question.options)
            .selectinload(QuestionOption.translations),
        )
        .order_by(Submission.created_at.desc())
    )
    if survey_id: q = q.where(Submission.survey_id == survey_id)
    if org_id:    q = q.where(Submission.organization_id == org_id)
    if role:      q = q.where(Submission.respondent_role == role)
    if status:    q = q.where(Submission.status == status)
    return list(db.execute(q.limit(10_000)).scalars().unique().all())


def _org_name(s: Submission) -> str:
    return (s.organization.name_en if s.organization else None) or s.org_name_input or ""


# ── Column definitions ────────────────────────────────────────────────────────

SUMMARY_HEADERS = [
    "Submission ID", "Role", "Sector", "Org Size",
    "Language", "Status", "Submitted At",
]

def _summary_meta(s: Submission) -> list:
    return [
        s.id, s.respondent_role, s.sector or "",
        s.org_size or "", s.language_used,
        str(s.status),
        s.submitted_at.isoformat() if s.submitted_at else "",
    ]

DETAIL_HEADERS = [
    "Submission ID", "Organization", "Respondent Name", "Role", "Email",
    "Sector", "Org Size", "Language", "Status", "Submitted At",
]

def _detail_meta(s: Submission) -> list:
    return [
        s.id, _org_name(s), s.respondent_name or "",
        s.respondent_role, s.respondent_email or "",
        s.sector or "", s.org_size or "", s.language_used,
        str(s.status),
        s.submitted_at.isoformat() if s.submitted_at else "",
    ]


# ── Intro questions ───────────────────────────────────────────────────────────

def _get_intro_questions(submissions: list[Submission]) -> list[Question]:
    seen: set[int] = set()
    intro_qs: list[Question] = []
    for s in submissions:
        for ans in s.answers:
            q = ans.question
            if q and getattr(q, "is_intro", False) and q.id not in seen:
                seen.add(q.id)
                intro_qs.append(q)
    return sorted(intro_qs, key=lambda q: q.display_order)

def _intro_pair(ans_map: dict, q: Question) -> list[str]:
    ans = ans_map.get(q.id)
    return [_answer_text(ans, "en"), _answer_text(ans, "ar")] if ans else ["", ""]


# ── XLSX sheet builder (reused per survey group) ──────────────────────────────

def _write_summary_sheet(ws, submissions: list[Submission], style_header_fn) -> None:
    intro_qs = _get_intro_questions(submissions)
    intro_col_hdrs: list[str] = []
    for iq in intro_qs:
        label = (_t(iq.translations, "en") or iq.question_key)[:50]
        intro_col_hdrs += [f"{label} (EN)", f"{label} (AR)"]

    style_header_fn(ws, SUMMARY_HEADERS + intro_col_hdrs)

    for s in submissions:
        ans_map = {a.question_id: a for a in s.answers}
        intro_vals: list[str] = []
        for iq in intro_qs:
            intro_vals += _intro_pair(ans_map, iq)
        ws.append(_summary_meta(s) + intro_vals)


def _write_answers_sheet(ws, submissions: list[Submission], style_header_fn) -> None:
    style_header_fn(ws, SUMMARY_HEADERS + [
        "Intro?", "Q#", "Question Key",
        "Question (EN)", "Question (AR)",
        "Answer (EN)", "Answer (AR)",
    ])
    for s in submissions:
        meta = _summary_meta(s)
        for i, ans in enumerate(
            sorted(s.answers, key=lambda a: a.question.display_order if a.question else 0),
            start=1,
        ):
            q = ans.question
            if not q:
                continue
            ws.append(meta + [
                "Yes" if getattr(q, "is_intro", False) else "No",
                i, q.question_key,
                _t(q.translations, "en"), _t(q.translations, "ar"),
                _answer_text(ans, "en"), _answer_text(ans, "ar"),
            ])


# ── Export Service ────────────────────────────────────────────────────────────

class ExportService:
    def __init__(self, db: Session):
        self.db = db

    def export_csv(
        self,
        survey_id: int | None = None,
        org_id:    int | None = None,
        role:      str | None = None,
        status:    str | None = None,
    ) -> bytes:
        submissions = _load_submissions(self.db, survey_id, org_id, role, status)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(DETAIL_HEADERS + [
            "Q#", "Question Key",
            "Question (EN)", "Question (AR)",
            "Answer (EN)", "Answer (AR)",
        ])

        for s in submissions:
            meta = _detail_meta(s)
            sorted_answers = sorted(
                s.answers,
                key=lambda a: a.question.display_order if a.question else 0,
            )
            if not sorted_answers:
                writer.writerow(meta + ["", "", "", "", "", ""])
                continue
            for i, ans in enumerate(sorted_answers, start=1):
                q = ans.question
                if not q:
                    continue
                writer.writerow(meta + [
                    i, q.question_key,
                    _t(q.translations, "en"), _t(q.translations, "ar"),
                    _answer_text(ans, "en"), _answer_text(ans, "ar"),
                ])

        return output.getvalue().encode("utf-8-sig")

    def export_xlsx(
        self,
        survey_id: int | None = None,
        org_id:    int | None = None,
        role:      str | None = None,
        status:    str | None = None,
    ) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        submissions = _load_submissions(self.db, survey_id, org_id, role, status)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default blank sheet

        header_fill = PatternFill("solid", fgColor="1B3A5C")
        header_font = Font(color="FFFFFF", bold=True)
        rtl_align   = Alignment(horizontal="right", readingOrder=2)
        center      = Alignment(horizontal="center")

        def style_header(ws, headers: list[str]) -> None:
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

        # Group submissions by survey slug (or id if no slug)
        survey_groups: dict[str, list[Submission]] = {}
        for s in submissions:
            key = (s.survey.slug if s.survey else None) or str(s.survey_id)
            survey_groups.setdefault(key, []).append(s)

        # One pair of sheets per survey
        for survey_slug, group in survey_groups.items():
            # Sheet names max 31 chars in Excel
            prefix = survey_slug[:26]
            ws_summary = wb.create_sheet(f"{prefix} - Sum")
            ws_answers  = wb.create_sheet(f"{prefix} - Ans")
            _write_summary_sheet(ws_summary, group, style_header)
            _write_answers_sheet(ws_answers,  group, style_header)

        # Apply RTL alignment to (AR) columns + auto-width
        for ws in wb.worksheets:
            ar_cols: set[int] = set()
            for cell in ws[1]:
                if cell.value and "(AR)" in str(cell.value):
                    ar_cols.add(cell.column)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column in ar_cols and cell.value:
                        cell.alignment = rtl_align
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
